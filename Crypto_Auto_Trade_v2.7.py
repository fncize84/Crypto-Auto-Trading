#----------------------------------------------------------------------------------------------------------------------
# Cryptocurrency Auto Trade
# Larry Williams Volatility Breakout Strategy + Moving Average
# Ver 2.5   거래종목 선정방식 변경 : 거래대금 상위 10개 종목 → 지정 종목 리스트
#           매매시간 변경 : 09:00 ~ 05:00 → 09:00 ~ 08:00
# Ver 2.6   투자 시작일, 최초 투자 금액 설정 방법 변경
# Ver 2.7   일일 매매 기준 시간 변경
#           매수 결과 메세지 수정 (Target Price 오류 수정)
#           매도 결과 메세지 수정 (수익/손실 금액 추가)
#           손절매 반영
#----------------------------------------------------------------------------------------------------------------------

import pyupbit
import time
import datetime
import schedule
import pandas as pd
import openpyxl
import requests
import matplotlib.pyplot as plt
from telegram_MSG import post_message, post_photo
from upbit_trade_record import order_record, trade_record, get_win_rate, cal_CAGR

program_version = 'Cyptocurrency Trade Bot v2.7'
strategy = "Larry's Volatility Breakout System"

key_path = '/home/ubuntu/upbit_key.txt'
file_path = '/home/ubuntu/trading_record.xlsx'

# key_path = 'upbit_key.txt'
# file_path = 'trading_record.xlsx'

k = 0.5
coin_nums = 5   # 일일 트레이딩 종목 수
stop_loss = 0.05   # 손절매 기준
tickers = ['KRW-BTC', 'KRW-ETH', 'KRW-XRP', 'KRW-SOL', 'KRW-AXS', 'KRW-MANA']

# Load account
with open(key_path) as f:
    lines = f.readlines()
    access = lines[0].strip()
    secret = lines[1].strip()
    upbit = pyupbit.Upbit(access, secret)

'''현재 보유 자산 현황 확인 / 초기투자금액 및 투자 시작일 설정 / 프로그램 실행시 1회만 실행'''
def trade_start_check():
    holdings = {ticker:False for ticker in tickers}
    balances = upbit.get_balances()
    for balance in balances:
        if balance['currency'] == "KRW":
            continue
        if float(balance['balance']) > 0:
            ticker = "KRW-" + str(balance['currency'])
            holdings[ticker] = True

    total_balance = upbit.get_balance("KRW") + upbit.get_amount('ALL')
    record_data = [datetime.date.today(), total_balance, pyupbit.get_ohlcv("KRW-BTC", count=2).iloc[-2]['close']]
    hold_list = []
    for ticker in tickers:
        if holdings[ticker] is True:
            hold = return_coin_name(ticker)
            hold_list.append(hold)
    hold_coins = ", ".join(hold_list) if len(hold_list) > 0 else "None"
    
    try:
        df = pd.read_excel(file_path, sheet_name = 'return_record')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'return_record'
        ws.append(['Date','total_assets','BTC'])
        ws.append(record_data)
        wb.save(file_path)
        df = pd.read_excel(file_path, sheet_name = 'return_record')

    T_start_date = df.iloc[0]['Date']
    T_start_date = datetime.date(T_start_date.year,T_start_date.month,T_start_date.day)
    balance_init = df.iloc[0]['total_assets']

    post_message('-'*35 + '\n' + program_version + '\n' + strategy +
                '\n' + 'Running Program on ' + str(datetime.datetime.now())[:16] + '\n' + '-'*35)
    post_message("Total Assets : {:0,.0f}won".format(total_balance)
                + '\n' + "Cash : {:0,.0f}won".format(upbit.get_balance("KRW"))
                + '\n' + "Holding Coins : " + hold_coins)
    return holdings, T_start_date, balance_init

def get_balance_unit(tickers):
    units = {ticker:0 for ticker in tickers}
    balances = upbit.get_balances()
    for balance in balances:
        if balance['currency'] == "KRW":
            continue
        ticker = "KRW-" + str(balance['currency'])
        unit = float(balance['balance'])
        units[ticker] = unit
    return units

def get_target_price(ticker):
    df = pyupbit.get_ohlcv(ticker, interval="day", count=10)
    break_out = df.iloc[-1]['open'] + (df.iloc[-2]['high'] - df.iloc[-2]['low']) * k
    ma5 = df['close'].rolling(5).mean().iloc[-2]
    return break_out, ma5

def inquiry_target_price(tickers):
    targets = {}
    ma5s = {}
    for t in tickers:
        targets[t], ma5s[t] = get_target_price(t)
    return targets, ma5s

def get_high_price(ticker):
    df = pyupbit.get_ohlcv(ticker, count=2)
    high_price = df['high'].iloc[-1]
    return high_price

def return_coin_name(ticker):
    url = "https://api.upbit.com/v1/market/all"
    coinname = requests.get(url)
    coinname = coinname.json()
    coinname_df = pd.DataFrame(coinname).set_index("market")
    coin_name = coinname_df.loc[ticker, "english_name"]
    return coin_name

'''매수 / 매도 / 셋업 시간 갱신 -  매일 09:01 실행'''
def trade_time_setup():
    start_time = pyupbit.get_ohlcv("KRW-BTC").index[-1]
    end_time = start_time + datetime.timedelta(hours=23)
    tomorrow = start_time + datetime.timedelta(days=1)
    sell_time = tomorrow - datetime.timedelta(minutes=5)
    return start_time, end_time, tomorrow, sell_time

'''매매 조건 갱신 -  매일 09:01 실행'''
def trade_setup():
    buy_amt = (upbit.get_balance("KRW") + upbit.get_amount('ALL')) / coin_nums
    targets, ma5s = inquiry_target_price(tickers)
    for ticker in tickers:
        target = targets[ticker]
        ma5 = ma5s[ticker]
        current = pyupbit.get_current_price(ticker)
        t = return_coin_name(ticker)
        post_message(t 
                    + '\n' + "Target Price : {:0,.0f}won".format(target)
                    + "\n" + "5 Days MA : {:0,.0f}won".format(ma5)
                    + "\n" + "Current Price : {:0,.0f}won".format(current))
    return buy_amt, targets, ma5s

def daily_report():
    balance_now = upbit.get_balance("KRW") + upbit.get_amount('ALL')
    total_count, win_count, win_rate, FnL_ratio= get_win_rate(file_path)
    t_delta = (datetime.date.today() - T_start_date).days
    FnL = (balance_now / balance_init - 1) * 100

    BTC_close = pyupbit.get_ohlcv("KRW-BTC", count=2).iloc[-2]['close']
    record_data = [datetime.date.today(),balance_now,BTC_close]

    try:
        wb = openpyxl.load_workbook(file_path)
        try:
            ws = wb['return_record']
        except KeyError:
            wb.create_sheet('return_record')
            ws = wb['return_record']
            ws.append(['Date','total_assets','BTC'])
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'return_record'
        ws.append(['Date','total_assets','BTC'])
    ws.append(record_data)
    wb.save(file_path)

    df = pd.read_excel(file_path, engine = "openpyxl", sheet_name="return_record")
    df['My_Port'] = df['total_assets'] / df['total_assets'].iloc[0] * 100
    df['BTC_INDEX'] = df['BTC'] / df['BTC'].iloc[0] * 100
    df['DD'] = (df['My_Port'] - df['My_Port'].cummax()) / df['My_Port'].cummax() * 100
    print(df)

    plt.figure(figsize=(12,8))
    
    plt.subplot(2, 1, 1)
    plt.plot(df['Date'], df['My_Port'], label = 'My Strategy')
    plt.plot(df['Date'], df['BTC_INDEX'], label = 'BTC buy and hold')
    plt.gca().spines['right'].set_visible(False)
    plt.gca().spines['left'].set_visible(False)
    plt.gca().grid(axis = 'y')
    plt.ylabel('Daily Return')
    plt.legend()

    plt.subplot(2, 1, 2)
    plt.plot(df['Date'], df['DD'], color = 'gray')
    plt.gca().spines['right'].set_visible(False)
    plt.gca().spines['left'].set_visible(False)
    plt.gca().spines['top'].set_visible(False)
    plt.gca().axes.xaxis.set_visible(False)
    plt.gca().grid(axis = 'y')
    plt.gca().fill_between(df['Date'], 0, df['DD'], color = 'lightgray')
    plt.ylabel('Draw Down Rate(%)')

    plt.savefig('daily_summary.png')
    save_fig = 'daily_summary.png'

    post_message("Trade Result summary : "
                + '\n' + "Total Assets : {:0,.0f}won".format(balance_now)
                + "\n" + "Total Returns : {:0,.1f}%".format(FnL)
                + "\n" + "Trading Period : {:0,.0f}days".format(t_delta)
                + "\n" + "CAGR : {:0,.1f}%".format(cal_CAGR(FnL,T_start_date))
                + "\n" + "Win Rate : {:0,.1f}%".format(win_rate) + "(" + str(win_count) + '/' + str(total_count) + ')'
                + "\n" + "F & L Ratio : {:0,.2f}".format(FnL_ratio))
    post_photo(save_fig)

holdings, T_start_date, balance_init = trade_start_check()
start_time, end_time, tomorrow, sell_time = trade_time_setup()
buy_amt, targets, ma5s = trade_setup()

post_message("Auto Trade Start...")
schedule.every().day.at("09:00").do(daily_report)

while True:
    try:
        schedule.run_pending()
        now = datetime.datetime.now()
        # 매수시간 : today 09:00:00 ~ tomorrow 09:00:00
        if start_time < now < sell_time:
            for ticker in tickers:
                price = pyupbit.get_current_price(ticker)
                high_price = get_high_price(ticker)
                krw = upbit.get_balance("KRW")
                target_price = targets[ticker] if targets[ticker] > ma5s[ticker] else ma5s[ticker]
                stop_loss_price = upbit.get_avg_buy_price(ticker) * (1 - stop_loss)

                # 매수 조건 
                cond_1 = target_price <= price < target_price * 1.005
                cond_2 = high_price <= price * 1.02
                cond_3 = holdings[ticker] is False
                cond_4 = krw > 5000
                cond_5 = start_time < now < end_time

                if cond_1 and cond_2 and cond_3 and cond_4 and cond_5:
                    buy_result = upbit.buy_market_order(ticker, buy_amt*0.9995)
                    order_record(buy_result, file_path)
                    avr_price = upbit.get_avg_buy_price(ticker)
                    slip = (1-avr_price/target_price) * 100
                    post_message(return_coin_name(ticker) + " - Target Price Breakout.."
                                + "\n" + "Target Price : {:0,.0f}won".format(target_price)
                                + "\n" + "Buy Price : {:0,.0f}won".format(avr_price)
                                + "\n" + "Slippage : {:0,.1f}%".format(slip))
                    holdings[ticker] = True

                if holdings[ticker] and price < stop_loss_price:
                    balance = upbit.get_balance(ticker)
                    sell_result = upbit.sell_market_order(ticker, balance)
                    order_record(sell_result, file_path)
                    fnl, win_rate, f_amt, l_amt = trade_record(ticker, file_path, access, secret)
                    fnl_amt = f_amt if win_rate == 1 else l_amt
                    post_message(return_coin_name(ticker) + "  Sell    " 
                                + "\n" + "Rate of Return : {:0,.2f}%".format(fnl)
                                + "\n" + "F & L Amount : {:0,.0f}won".format(fnl_amt))

        #매도시간 : tomorrow 08:55:00 ~ 09:00:00
        #보유종목 전량 시장가 매도 주문 / 미체결주문 취소 / 보유종목 리스트 초기화
        if sell_time < now < tomorrow: 
            units = get_balance_unit(tickers)
            for ticker in tickers:
                unit = units.get(ticker, 0)
                if unit > 0:
                    sell_result = upbit.sell_market_order(ticker, unit)
                    order_record(sell_result, file_path)
                    fnl, win_rate, f_amt, l_amt = trade_record(ticker, file_path, access, secret)
                    fnl_amt = f_amt if win_rate == 1 else l_amt
                    post_message(return_coin_name(ticker) + "  Sell    " 
                                + "\n" + "Rate of Return : {:0,.2f}%".format(fnl)
                                + "\n" + "F & L Amount : {:0,.0f}won".format(fnl_amt))
            holdings = {ticker:False for ticker in tickers}
        # 매매조건 갱신 : tomorrw 09:01:00 ~ 09:01:10
        if tomorrow < now:
            start_time, end_time, tomorrow, sell_time = trade_time_setup()
            buy_amt, targets, ma5s = trade_setup()
        time.sleep(1)
    except Exception as e:
        print(e)
        post_message(str(e))
        time.sleep(1)