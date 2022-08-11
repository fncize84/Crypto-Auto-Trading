import openpyxl
import pyupbit
import datetime
import pandas as pd

# 매수/매도 주문내역 엑셀 저장
def order_record(result, file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb['order_record']
        
    except KeyError:
        wb.create_sheet('order_record')
        ws = wb['order_record']
   
    ws.append([str(result)])
    wb.save(file_path)

# 매매결과 엑셀저장_수익률 계산
def trade_record(ticker, file_path, access, secret):
    buy_c = 0
    sell_c = 0
    upbit = pyupbit.Upbit(access, secret)
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb['order_record']
        for a in range(10):
            cell_no = 'A'+str((len(ws['A'])-a))
            order_Data = eval(ws[cell_no].value)
            if order_Data['market'] == ticker:
                if order_Data['side'] == 'bid' and buy_c == 0 :
                    buy_c = float(order_Data['price'])
                elif order_Data['side'] == 'ask' and sell_c == 0:
                    sell_uuid = order_Data['uuid']
                    sell_order = upbit.get_order(sell_uuid)
                    paid_fee = float(sell_order['paid_fee'])
                    last_trade = sell_order['trades']
                    funds = 0
                    for t in last_trade:
                        funds += float(t['funds'])        
                    sell_c = funds - paid_fee
            if buy_c > 0 and sell_c > 0:
                break
        fnl = (sell_c/buy_c -1) * 100
        win_rate = 1 if fnl >= 0 else 0
        f_amt = sell_c - buy_c if win_rate == 1 else 0
        l_amt = buy_c - sell_c if win_rate == 0 else 0
        record_data = [datetime.date.today(),ticker,buy_c,sell_c,fnl,win_rate,f_amt,l_amt]
        try:
            ws1 = wb['trade_record']
        except KeyError:
            wb.create_sheet('trade_record')
            ws1 = wb['trade_record']
            ws1.append(['날짜','종목명','매입금액','매도금액','수익률','수익/손실','수익금액','손실금액'])
        ws1.append(record_data)
        wb.save(file_path)
    except Exception:
        fnl = 0        
    return fnl, win_rate, f_amt, l_amt

# 매매 수익률 계산
def get_win_rate(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb['trade_record']
        data = ws.values
        columns = next(data)[0:]
        DF = pd.DataFrame(data, columns = columns)
        total_count = DF['수익/손실'].count()
        try:
            win_count = DF['수익/손실'].value_counts()[1]
        except KeyError:
            win_count = 0
        win_rate = (win_count / total_count) * 100
        f_avr = DF['수익금액'].sum() / win_count
        l_avr = DF['손실금액'].sum() / (total_count - win_count)
        fnl_ratio = f_avr / l_avr
    except Exception:
        total_count = win_count = win_rate = fnl_ratio = 0
    return total_count, win_count, win_rate, fnl_ratio

# CAGR 계산
def cal_CAGR(FnL, T_start_date):
    now_date = datetime.date.today()
    t_range = (now_date - T_start_date).days /365
    CAGR = (((FnL / 100 + 1) ** (1/t_range))- 1)*100
    return CAGR
