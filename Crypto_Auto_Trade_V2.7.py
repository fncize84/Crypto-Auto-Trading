#----------------------------------------------------------------------------------------------------------------------
# Cryptocurrency Auto Trade
# Larry Williams Volatility Breakout Strategy + Moving Average
# Ver 2.5   2022. 03. 01
#           거래종목 선정방식 변경 : 거래대금 상위 10개 종목 → 지정 종목 리스트
#           매매시간 변경 : 09:00 ~ 05:00 → 09:00 ~ 08:00
# 
# Ver 2.6   2022. 03. 04 
#           Stop Loss 변경 : 3% → 5%
#           투자원금, 투자시작일 변수설정 변경
#
# Ver 2.7   2022. 03. 08
#           매수, 매도 주문 방식 변경 : 시장가 → 지정가
#----------------------------------------------------------------------------------------------------------------------

import pyupbit
import time
import datetime
import schedule
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import trade_utility as ut
from Upb_Get_Data import get_target_price, get_high_price, get_current_price, get_ma5, get_start_time, return_coin_name, get_ticker_list, get_balance
from upbit_trade_record import order_record, trade_record, get_win_rate, cal_CAGR

Debug_ = False       # True: 매매 API 호출 안됨, False: 실제로 매매 API 호출

program_version = 'Cyptocurrency Trade Bot v2.7'
strategy = "Larry's Volatility Breakout System"
key_path = '/home/ubuntu/upbit_key.txt' if Debug_ is False else 'upbit_key.txt'
file_path = '/home/ubuntu/trading_record.xlsx' if Debug_ is False else 'trading_record.xlsx'

k = 0.5
trading_size = 5   # 일일 트레이딩 종목 수
stop_loss = 0.05   # 손절매 기준
ticker = ['KRW-BTC', 'KRW-ETH', 'KRW-XRP', 'KRW-SOL', 'KRW-AXS', 'KRW-MANA']

# Load account
with open(key_path) as f:
    lines = f.readlines()
    access = lines[0].strip()
    secret = lines[1].strip()
    upbit = pyupbit.Upbit(access, secret)

def trade_start_check():
    '''현재 보유중인 암호화폐 리스트 확인하는 함수 / 프로그램 실행시 1회만 실행'''
    balance_now = upbit.get_balance("KRW") + upbit.get_amount('ALL')
    global buy_list
    holdings = []
    if len(buy_list) > 0:
        for b in buy_list:
            holding = return_coin_name(b) 
            holdings.append(holding)
        hold_list = ", ".join(holdings)
    else:
        hold_list = "None"

    record_data = [datetime.date.today(),balance_now,pyupbit.get_ohlcv("KRW-BTC", count=2).iloc[-2]['close']]
    try:
        df = pd.read_excel(file_path, sheet_name = 'return_record')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'return_record'
        ws.append(['Date','total_assets','BTC'])
        ws.append(record_data)
        wb.save(file_path)
        df = pd.read_excel(file_path, sheet_name = 'return_record')

    global T_start_date
    global balance_init
    T_start_date = df.iloc[0]['Date']
    T_start_date = datetime.date(T_start_date.year,T_start_date.month,T_start_date.day)
    balance_init = df.iloc[0]['total_assets']

    ut.post_message('-'*35 + '\n' + program_version + '\n' + strategy +
                    '\n' + 'Running Program on ' + str(datetime.datetime.now())[:16] + '\n' + '-'*35)
    ut.post_message("Total Assets : {:0,.0f}won".format(balance_now)
                    + '\n' + "Cash : {:0,.0f}won".format(upbit.get_balance("KRW"))
                    + '\n' + "Holding Coins : " + hold_list)

def daily_report():
    global T_start_date
    global balance_init
    balance_now = upbit.get_balance("KRW") + upbit.get_amount('ALL')
    total_count, win_count, win_rate, FnL_ratio= get_win_rate(file_path)
    t_delta = (datetime.date.today() - T_start_date).days
    FnL = (balance_now / balance_init - 1) * 100

    BTC_close = pyupbit.get_ohlcv("KRW-BTC", count=2).iloc[-2]['close']
    record_data = [datetime.date.today(),balance_now,BTC_close]

    try:
        wb = openpyxl.load_workbook(file_path)
        try:
            ws = wb['return_record']
        except KeyError:
            wb.create_sheet('return_record')
            ws = wb['return_record']
            ws.append(['Date','total_assets','BTC'])
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'return_record'
        ws.append(['Date','total_assets','BTC'])
    ws.append(record_data)
    wb.save(file_path)

    df = pd.read_excel(file_path, engine = "openpyxl", sheet_name="return_record")
    df['My_Port'] = df['total_assets'] / df['total_assets'].iloc[0] * 100
    df['BTC_INDEX'] = df['BTC'] / df['BTC'].iloc[0] * 100
    df['DD'] = (df['My_Port'] - df['My_Port'].cummax()) / df['My_Port'].cummax() * 100

    plt.figure(figsize=(12,8))
    
    plt.subplot(2, 1, 1)
    plt.plot(df['Date'], df['My_Port'], label = 'My Strategy')
    plt.plot(df['Date'], df['BTC_INDEX'], label = 'BTC buy and hold')
    plt.gca().spines['right'].set_visible(False)
    plt.gca().spines['left'].set_visible(False)
    plt.gca().grid(axis = 'y')
    plt.ylabel('Daily Return')
    plt.legend()

    plt.subplot(2, 1, 2)
    plt.plot(df['Date'], df['DD'], color = 'gray')
    plt.gca().spines['right'].set_visible(False)
    plt.gca().spines['left'].set_visible(False)
    plt.gca().spines['top'].set_visible(False)
    plt.gca().axes.xaxis.set_visible(False)
    plt.gca().grid(axis = 'y')
    plt.gca().fill_between(df['Date'], 0, df['DD'], color = 'lightgray')
    plt.ylabel('Draw Down Rate(%)')

    plt.savefig('daily_summary.png')
    save_fig = 'daily_summary.png'

    ut.post_message("Trade Result summary : "
                    + '\n' + "Total Assets : {:0,.0f}won".format(balance_now)
                    + "\n" + "Total Returns : {:0,.1f}%".format(FnL)
                    + "\n" + "Trading Period : {:0,.0f}days".format(t_delta)
                    + "\n" + "CAGR : {:0,.1f}%".format(cal_CAGR(FnL,T_start_date))
                    + "\n" + "Win Rate : {:0,.1f}%".format(win_rate) + "(" + str(win_count) + '/' + str(total_count) + ')'
                    + "\n" + "F & L Ratio : {:0,.2f}".format(FnL_ratio))
    ut.post_photo(save_fig)

def trade_setup():
    '''일일 트레이딩 데이터 셋업 / 매수종목 리스트 초기화 / 매일 09:00 실행'''
    global buy_list
    buy_list = get_balance(access, secret)

trade_setup()
trade_start_check()
ut.post_message("Auto Trade Start...")

set_time = "09:00"
schedule.every().day.at(set_time).do(daily_report)
schedule.every().day.at(set_time).do(trade_setup)

# 금일 09:05 ~ 익일 08:00 : 매수 진행      목표가 돌파가 너무 늦은 시간에 이뤄지는 경우 바로 매도가 이뤄질수 있어 매매 시간은 23시간만 진행
# 익일 08:55  :  보유 코인 일괄 매도
# 익일 09:00  : 매매 세팅 

while True:
    try:
        schedule.run_pending()
        buy_CNT = len(buy_list)
        krw = upbit.get_balance("KRW")
        buy_amount = 0 if buy_CNT == trading_size else krw/(trading_size-buy_CNT)

        now = datetime.datetime.now()
        start_time = get_start_time()
        end_time = start_time + datetime.timedelta(hours=23)
        sell_time = start_time + datetime.timedelta(days=1) - datetime.timedelta(minutes=5)

        if start_time < now < sell_time:
            for t in ticker:
                if t not in buy_list:
                    target_price = get_target_price(t,k)
                    ma5 = get_ma5(t)
                    high_price = get_high_price(t)
                    current_price = get_current_price(t)
                    stop_loss_price = upbit.get_avg_buy_price(t) * (1 - stop_loss)
                    balance = upbit.get_balance(t)
                    
                    cond_1 = target_price <= current_price and ma5 <= current_price
                    cond_2 = high_price < current_price * 1.02
                    cond_3 = krw > 5000
                    cond_4 = start_time < now < end_time

                    if cond_1 and cond_2 and cond_3 and cond_4:
                        if Debug_ is False:
                            orderbook = pyupbit.get_orderbook(t)[0]['orderbook_units'][0]
                            buy_price = int(orderbook['ask_price'])
                            buy_unit = buy_amount*0.9995 / buy_price
                            buy_result = upbit.buy_limit_order(t, buy_price, buy_unit)
                            buy_list.append(t)
                        else:
                            print("BUY API CALLED : ", t)

                else:

                    order_record(buy_result, file_path)
                    time.sleep(1)
                    avr_price = upbit.get_avg_buy_price(t)
                    slip = (1-avr_price/target_price) * 100
                    ut.post_message(return_coin_name(t) + " - Target Price Breakout.."
                                    + "\n" + "Target Price : {:0,.0f}won".format(target_price)
                                    + "\n" + "Buy Price : {:0,.0f}won".format(avr_price)
                                    + "\n" + "Slippage : {:0,.1f}%".format(slip))

                    if balance > 0 and current_price < stop_loss_price:
                        if Debug_ is False:
                            sell_result = upbit.sell_market_order(t, balance)
                            order_record(sell_result, file_path)
                            ut.post_message(return_coin_name(t) + "  Sell    " 
                                            + "\n" + "Rate of Return : {:0,.2f}%".format(trade_record(t, file_path, access, secret)))
                        else:
                            print("SELL_STOP LOSS API CALLED : ", t)

        elif sell_time < now:
            for t in ticker:
                balance = upbit.get_balance(t)
                if balance > 0:
                    if Debug_ is False:
                        sell_result = upbit.sell_market_order(t, balance)
                        order_record(sell_result, file_path)
                        ut.post_message(return_coin_name(t) + "  Sell    " 
                                        + "\n" + "Rate of Return: {:0,.2f}%".format(trade_record(t, file_path, access, secret)))
                    else:
                        print("SELL API CALLED : ", t)
        time.sleep(1)
    except Exception as e:
        print(e)
        ut.post_message(str(e))
        time.sleep(1)