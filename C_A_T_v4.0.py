import pyupbit
import time
import datetime
import telegram
import schedule
import pandas as pd
import openpyxl
from Upb_Get_Data import get_target_price, get_current_price, get_ma5, get_start_time, return_coin_name

program_version = 'Auto Trade v4.0_2022-02-07'

access = "2nUWiu0PnEtvlUlpiq5PJGc0b1N4ix5fPox6KRYA"
secret = "AmSsPXpxvx12UEZsxBIIPyLU0ZVsDiMl2cSLtX5X"

token = "5094155373:AAGwbZOBTw990tvU6TIdHWilsHP7R95T-qM"
chat_id = "5033041863"

def post_message(chat_id, message):
	bot = telegram.Bot(token)
	bot.sendMessage(chat_id=chat_id, text=message)

def get_balance():
    balances = upbit.get_balances()
    ticker = []
    for b in balances:
        if b['currency'] != "KRW":
            t = "KRW-"+str(b['currency'])
            ticker.append(t)
    return(ticker)

def get_ticker_list():
    """ 최근 5일 평균 거래대금 상위 10개 종목 선정 """    
    Ticker = pyupbit.get_tickers(fiat="KRW")
    DF = pd.DataFrame(columns = ["ticker","trading amount"])
    for t in Ticker:
        df = pyupbit.get_ohlcv(t, period=1, interval="day", count=6) 
        df = df.iloc[:-1]
        mean_amount = df["value"].mean()   
        DF.loc[len(DF)]=[t,mean_amount]
        time.sleep(0.1)
    DF = DF.sort_values(by=['trading amount'], ascending=False, axis=0).iloc[:10]
    ticker_list = DF["ticker"].tolist()
    return ticker_list

def get_sell_amount(uuid):
    """ 매도금액 조회(수수료 포함) """
    sell_order = upbit.get_order(uuid)
    paid_fee = float(sell_order['paid_fee'])
    last_trade = sell_order['trades']
    funds = 0
    for t in last_trade:
        funds += float(t['funds'])        
    sell_amount = funds - paid_fee
    return sell_amount

def order_record(result):
    """ 매수/매도 주문내역 엑셀 저장 """
    try:
        wb = openpyxl.load_workbook('trading_record.xlsx')
        ws = wb['order_record']
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'order_record'    
    ws.append([str(result)])
    wb.save('trading_record.xlsx')

def trade_record(ticker):
    """ 매매결과 엑셀저장_수익률 계산 """
    buy_c = 0
    sell_c = 0
    try:
        wb = openpyxl.load_workbook('trading_record.xlsx')
        ws = wb['order_record']
        for a in range(10):
            cell_no = 'A'+str((len(ws['A'])-a))
            order_Data = eval(ws[cell_no].value)
            if order_Data['market'] == ticker:
                if order_Data['side'] == 'bid' and buy_c == 0 :
                    buy_c = float(order_Data['price'])
                elif order_Data['side'] == 'ask' and sell_c == 0:
                    sell_uuid = order_Data['uuid']
                    sell_c = get_sell_amount(sell_uuid)
            if buy_c > 0 and sell_c > 0:
                break
        fnl = (sell_c/buy_c -1)*100
        win_rate = 1 if fnl >= 0 else 0
        record_data = [datetime.date.today(),ticker,buy_c,sell_c,fnl,win_rate]
        try:
            ws1 = wb['trade_record']
        except KeyError:
            wb.create_sheet('trade_record')
            ws1 = wb['trade_record']
            ws1.append(['날짜','종목명','매입금액','매도금액','수익률','수익/손실'])
        ws1.append(record_data)
        wb.save('trading_record.xlsx')
    except Exception:
        fnl = 0        
    return fnl

def get_win_rate():
    try:
        wb = openpyxl.load_workbook('trading_record.xlsx')
        ws = wb['trade_record']
        data = ws.values
        columns = next(data)[0:]
        DF = pd.DataFrame(data, columns = columns)
        total_count = DF['수익/손실'].count()
        win_count = DF['수익/손실'].value_counts()[1]
        win_rate = (win_count / total_count) * 100
    except Exception:
        total_count = win_count = win_rate = 0
    return total_count, win_count, win_rate

def cal_CAGR(FnL, T_start_date):
    now_date = datetime.date.today()
    t_range = (now_date - T_start_date).days /365
    CAGR = (((FnL / 100 + 1) ** (1/t_range))- 1)*100
    return CAGR

def auto_trade_start():
    post_message(chat_id,'-----------------------------')
    post_message(chat_id,program_version)
    post_message(chat_id,str(datetime.date.today()))
    total_count, win_count, win_rate = get_win_rate()
    balance_now = upbit.get_balance("KRW")+upbit.get_amount('ALL')
    t_delta = (datetime.date.today() - T_start_date).days
    FnL = (balance_now / balance_init - 1)*100
    global ticker
    global buy_list
    ticker = get_ticker_list()
    buy_list = get_balance()
    post_message(chat_id,"보유종목")
    post_message(chat_id,buy_list)
    post_message(chat_id,"원화잔고 : {:0,.0f}원".format(upbit.get_balance("KRW")))
    post_message(chat_id, "총 평가금액 : {:0,.0f}원".format(balance_now))
    post_message(chat_id,"누적 수익률 : {:0,.1f}%".format(FnL) + "  총 투자일수 : {:0,.0f}일".format(t_delta))
    post_message(chat_id, "CAGR : {:0,.1f}%".format(cal_CAGR(FnL,T_start_date)) + "   승률 : {:0,.1f}%".format(win_rate) + "(" + str(win_count) + '/' + str(total_count) + ')') 
    today_list = []
    for t in ticker:
        today_list.append(return_coin_name(t))
    post_message(chat_id,"오늘의 매매 종목 : ")
    post_message(chat_id,today_list)
    post_message(chat_id, "자동매매 시작..")

upbit = pyupbit.Upbit(access, secret)

k = 0.5
trading_size = 5   # 일일 트레이딩 종목 수
stop_loss = 0.03   # 손절매 기준
balance_init = 1000000
T_start_date = datetime.date(2022,1,20)

# 자동매매 시작
auto_trade_start()
schedule.every().day.at("09:00").do(auto_trade_start)

while True:
    try:
        schedule.run_pending()
        count = len(upbit.get_balances())-1
        krw = upbit.get_balance("KRW")
        buy_amount = 0 if count == 5 else krw/(trading_size-count)
        now = datetime.datetime.now()
        start_time = get_start_time("KRW-BTC")
        end_time = start_time + datetime.timedelta(hours=18)
        sell_time = start_time + datetime.timedelta(days=1) - datetime.timedelta(minutes=10)

        if start_time < now < sell_time:
            for t in ticker:
                target_price = get_target_price(t)
                ma5 = get_ma5(t)
                current_price = get_current_price(t)
                balance = upbit.get_balance(t)
                stop_loss_price = upbit.get_avg_buy_price(t) * (1 - stop_loss)
                if target_price < current_price and ma5 < current_price and start_time < now < end_time:
                    if t not in buy_list and krw > 5000:
                        buy_result = upbit.buy_market_order(t, buy_amount*0.9995)
                        buy_list.append(t)
                        order_record(buy_result)
                        post_message(chat_id, return_coin_name(t) + "  목표가 돌파.. 매수")
                if balance > 0 and current_price < stop_loss_price:
                    sell_result = upbit.sell_market_order(t, balance)
                    order_record(sell_result)
                    post_message(chat_id, return_coin_name(t) + "  매도    " + "수익률 : {:0,.2f}%".format(trade_record(t)))

        elif sell_time < now:
            for t in ticker:
                balance = upbit.get_balance(t)
                if balance > 0:
                    sell_result = upbit.sell_market_order(t, balance)
                    order_record(sell_result)
                    post_message(chat_id, return_coin_name(t) + "  매도    " + "수익률 : {:0,.2f}%".format(trade_record(t)))
        time.sleep(1)

    except Exception as e:
        print(e)
        post_message(chat_id, str(e))
        time.sleep(1)
