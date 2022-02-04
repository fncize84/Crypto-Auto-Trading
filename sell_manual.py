import pyupbit
import datetime
import telegram
import pandas as pd
import requests
import openpyxl

access = "2nUWiu0PnEtvlUlpiq5PJGc0b1N4ix5fPox6KRYA"
secret = "AmSsPXpxvx12UEZsxBIIPyLU0ZVsDiMl2cSLtX5X"

token = "5094155373:AAGwbZOBTw990tvU6TIdHWilsHP7R95T-qM"
chat_id = "5033041863"
upbit = pyupbit.Upbit(access, secret)

def get_balance():
    balances = upbit.get_balances()
    ticker = []
    for b in balances:
        if b['currency'] != "KRW":
            t = "KRW-"+str(b['currency'])
            ticker.append(t)
    return(ticker)

def post_message(chat_id, message):
	bot = telegram.Bot(token)
	bot.sendMessage(chat_id=chat_id, text=message)

def order_record(result):
    """ 매수/매도 주문내역 엑셀 저장 """
    try:
        wb = openpyxl.load_workbook('trading_record.xlsx')
        ws = wb['order_record']
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'order_record'    
    ws.append([str(result)])
    wb.save('trading_record.xlsx')

def get_sell_amount(uuid):
    """ 매도금액 조회(수수료 포함) """
    sell_order = upbit.get_order(uuid)
    paid_fee = float(sell_order['paid_fee'])
    last_trade = sell_order['trades']
    funds = 0
    for t in last_trade:
        funds += float(t['funds'])        
    sell_amount = funds - paid_fee
    return sell_amount

def trade_record(ticker):
    """ 매매결과 엑셀저장_수익률 계산 """
    buy_c = 0
    sell_c = 0
    try:
        wb = openpyxl.load_workbook('trading_record.xlsx')
        ws = wb['order_record']
        for a in range(10):
            cell_no = 'A'+str((len(ws['A'])-a))
            order_Data = eval(ws[cell_no].value)
            if order_Data['market'] == ticker:
                if order_Data['side'] == 'bid' and buy_c == 0 :
                    buy_c = float(order_Data['price'])
                elif order_Data['side'] == 'ask' and sell_c == 0:
                    sell_uuid = order_Data['uuid']
                    sell_c = get_sell_amount(sell_uuid)
            if buy_c > 0 and sell_c > 0:
                break

        fnl = (sell_c/buy_c -1)*100
        win_rate = 1 if fnl >= 0 else 0   
        record_data = [datetime.date.today(),ticker,buy_c,sell_c,fnl,win_rate]

        try:
            ws1 = wb['trade_record']
        except KeyError:
            wb.create_sheet('trade_record')
            ws1 = wb['trade_record']
            ws1.append(['날짜','종목명','매입금액','매도금액','수익률','수익/손실'])
        ws1.append(record_data)
        wb.save('trading_record.xlsx')

    except Exception:
        fnl = 0        

    return fnl

def return_coin_name(ticker):
    url = "https://api.upbit.com/v1/market/all"
    coinname = requests.get(url)  # api 데이터 호출
    coinname = coinname.json()  # coinname으로 가져온 json 데이터를 list로 저장
    coinname_df = pd.DataFrame(coinname).set_index("market")
    coin_name = coinname_df.loc[ticker, "english_name"]
    return coin_name

ticker = get_balance()
for t in ticker:
    balance = upbit.get_balance(t)
    sell_result = upbit.sell_market_order(t, balance)
    order_record(sell_result)
    post_message(chat_id, return_coin_name(t) + "  매도    " + "수익률 : {:0,.2f}%".format(trade_record(t)))



